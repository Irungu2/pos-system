# ==========================================
# Standard Library Imports
# ==========================================
import base64
import io
import logging
from io import BytesIO

# ==========================================
# Third-Party Imports
# ==========================================
import barcode
import pandas as pd
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from reportlab.lib.pagesizes import mm
from reportlab.lib.units import mm as unit_mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

# ==========================================
# Django Imports
# ==========================================
from django.core.exceptions import PermissionDenied
from django.db import models, transaction
from django.db.models import Count, F, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_http_methods

# ==========================================
# Django REST Framework Imports
# ==========================================
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

# ==========================================
# Django Filters
# ==========================================
from django_filters.rest_framework import DjangoFilterBackend

# ==========================================
# Project Services
# ==========================================
from .services import (
    RestockDebugService,
    StoreStockService,
)

# ==========================================
# Project Utilities
# ==========================================
from .store_utils import get_current_store

# ==========================================
# Models
# ==========================================
from .models import (
    BulkRestock,
    BulkRestockItem,
    Category,
    Product,
    StockTransaction,
    StockTransfer,
    Store,
    StoreStock,
)

# ==========================================
# Serializers
# ==========================================
from .serializers import (
    BulkRestockSerializer,
    CategorySerializer,
    POSProductSerializer,
    ProductSerializer,
    StockTransactionSerializer,
    StockTransferCreateSerializer,
    StockTransferSerializer,
    StoreSerializer,
    StoreStockSerializer,
)
# =========================
# Logging Setup
# =========================
logger = logging.getLogger(__name__)
@method_decorator(csrf_exempt, name='dispatch')
class CategoryViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']

    def initial(self, request, *args, **kwargs):
        """Debug CSRF token"""
        super().initial(request, *args, **kwargs)
        
        logger.info(f"CSRF Debug - Request method: {request.method}")
        logger.info(f"CSRF Debug - Headers: {dict(request.headers)}")
        logger.info(f"CSRF Debug - CSRF_COOKIE: {request.META.get('CSRF_COOKIE')}")
        logger.info(f"CSRF Debug - HTTP_X_CSRFTOKEN: {request.META.get('HTTP_X_CSRFTOKEN')}")
        
        # For non-safe methods, check CSRF
        if request.method not in ('GET', 'HEAD', 'OPTIONS', 'TRACE'):
            csrf_token = request.META.get('HTTP_X_CSRFTOKEN')
            if csrf_token:
                logger.info(f"CSRF token received: {csrf_token[:20]}... (length: {len(csrf_token)})")
            else:
                logger.warning("No CSRF token in headers")

    def create(self, request, *args, **kwargs):
        logger.info(f"Create request - Data: {request.data}")
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        logger.info(f"Update request - Data: {request.data}")
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        logger.info(f"Delete request for category ID: {kwargs.get('pk')}")
        return super().destroy(request, *args, **kwargs)

    def perform_destroy(self, instance):
        if instance.products.exists():
            raise ValidationError(
                {'error': 'Cannot delete category with associated products'}
            )
        instance.delete()






@method_decorator(csrf_exempt, name='dispatch')
class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all().select_related("category")
    serializer_class = ProductSerializer

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["category", "is_active"]
    search_fields = ["name", "sku", "barcode", "description"]
    ordering_fields = ["name", "selling_price", "created_at"]
    ordering = ["name"]

    def get_queryset(self):
        user = self.request.user
        qs = Product.objects.all().select_related("category")
        
        # Get context from query parameter (default to 'inventory' for admins, 'pos' for others)
        context = self.request.query_params.get('context', '')
        store_mode = self.request.query_params.get('store_mode', '')
        
        # Determine if this is a POS request
        is_pos_request = context == 'pos' or store_mode == 'store'
        
        # 👑 ADMIN / MANAGER
        if user.is_superuser or user.role in ["admin", "manager"]:
            # For POS requests, filter by store (admin acting as cashier)
            if is_pos_request:
                try:
                    store = get_current_store(self.request)
                    if store:
                        return qs.filter(
                            storestock__store=store,
                            storestock__quantity__gt=0
                        ).distinct()
                except PermissionDenied:
                    pass
                return qs.none()
            
            # For inventory management, show ALL products
            return qs.distinct()

        # 👤 NORMAL USER (cashier, salesperson)
        # Always enforce store filtering for POS
        try:
            store = get_current_store(self.request)
            if store:
                return qs.filter(
                    storestock__store=store,
                    storestock__quantity__gt=0
                ).distinct()
        except PermissionDenied:
            pass
        
        # If no store selected or permission denied, return empty
        return qs.none()

    @action(detail=False, methods=["get"], url_path="low_stock")
    def low_stock(self, request):
        try:
            store = get_current_store(request)
            if not store:
                return Response({"error": "No store selected"}, status=400)
            
            # Get context to determine if admin wants store-specific or global view
            context = request.query_params.get('context', 'pos')
            user = request.user
            is_admin = user.is_superuser or user.role in ["admin", "manager"]
            
            if context == 'inventory' and is_admin:
                # Admin viewing low stock across ALL stores
                from django.db.models import Sum, Min, OuterRef, Subquery
                
                # Get products that are low in ANY store
                low_stock_products = Product.objects.filter(
                    storestock__quantity__lte=F('reorder_level')
                ).distinct()
                
                # Annotate with total stock across all stores
                low_stock_products = low_stock_products.annotate(
                    total_stock=Sum('storestock__quantity'),
                    min_stock=Min('storestock__quantity')
                )
                
                serializer = self.get_serializer(low_stock_products, many=True)
                return Response({
                    "context": "inventory",
                    "stores": "all",
                    "count": low_stock_products.count(),
                    "results": serializer.data
                })
            else:
                # Store-specific low stock (for POS or non-admin)
                qs = Product.objects.filter(
                    storestock__store=store,
                    storestock__quantity__lte=F("reorder_level")
                ).distinct()
                
                serializer = self.get_serializer(qs, many=True)
                return Response({
                    "context": "pos",
                    "store_id": store.id,
                    "store_name": store.name,
                    "count": qs.count(),
                    "results": serializer.data
                })
            
        except PermissionDenied as e:
            return Response({"error": str(e)}, status=403)

    @action(detail=False, methods=["get"], url_path="search_products")
    def search_products(self, request):
        print("\n" + "="*50)
        print("[PRODUCT API] search_products called")
        print("="*50)
        
        query = request.query_params.get("q", "")
        category = request.query_params.get("category")
        context = request.query_params.get('context', 'inventory')
        store_mode = request.query_params.get('store_mode', '')
        
        print(f"[PRODUCT API] Search query: '{query}'")
        print(f"[PRODUCT API] Category filter: '{category}'")
        print(f"[PRODUCT API] Context: '{context}'")
        print(f"[PRODUCT API] Store mode: '{store_mode}'")
        print(f"[PRODUCT API] User: {request.user.display_name} (Role: {request.user.role})")
        
        user = request.user
        qs = Product.objects.all().select_related("category")
        is_admin = user.is_superuser or user.role in ["admin", "manager"]
        
        print(f"\n[STEP 1] Initial queryset count: {qs.count()}")
        print(f"[STEP 1a] Is Admin: {is_admin}")
        
        # Determine filtering strategy
        force_store_filter = False
        is_pos_context = (context == 'pos' or store_mode == 'store')
        
        # POS context always forces store filtering with stock > 0
        if is_pos_context:
            print(f"\n[STEP 2] POS CONTEXT - enforcing store filter with stock > 0")
            force_store_filter = True
            show_only_in_stock = True  # ← CRITICAL: Always true for POS
        elif not is_admin:
            print(f"\n[STEP 2] NON-ADMIN USER - enforcing store filter")
            force_store_filter = True
            show_only_in_stock = True  # Non-admins also only see in-stock
        else:
            print(f"\n[STEP 2] ADMIN IN INVENTORY MODE - showing all products")
            force_store_filter = False
            show_only_in_stock = False
        
        # Apply store filter if needed
        if force_store_filter:
            try:
                store = get_current_store(request)
                if store:
                    print(f"[STEP 2a] Current store: {store.name} (ID: {store.id})")
                    print(f"[STEP 2b] Filtering products by store_id={store.id}")
                    print(f"[STEP 2c] Show only in-stock products: {show_only_in_stock}")
                    
                    # CRITICAL FIX: Always use quantity__gt=0 for POS
                    if show_only_in_stock:
                        qs = qs.filter(
                            storestock__store=store,
                            storestock__quantity__gt=0  # ← Only positive stock
                        ).distinct()
                        print(f"[STEP 2d] POS mode - only products with stock > 0")
                    else:
                        qs = qs.filter(storestock__store=store).distinct()
                        print(f"[STEP 2d] Inventory mode - all products in store")
                    
                    print(f"[STEP 2e] After store filter count: {qs.count()}")
                    
                    # Print products after filter for debugging
                    print(f"\n[PRODUCTS IN STORE {store.name} (stock > 0 only)]:")
                    for idx, p in enumerate(qs[:10], 1):
                        stock = StoreStock.objects.filter(store=store, product=p).first()
                        stock_qty = stock.quantity if stock else 0
                        print(f"  {idx}. {p.name} (ID: {p.id}) - Stock: {stock_qty}")
                else:
                    print(f"[STEP 2a] No store selected! Returning error")
                    return Response(
                        {"error": "No store selected. Please select a store first."}, 
                        status=400
                    )
            except PermissionDenied as e:
                print(f"[STEP 2 ERROR] Permission denied: {str(e)}")
                return Response({"error": str(e)}, status=403)
        else:
            print(f"\n[STEP 2] ADMIN INVENTORY MODE - showing ALL products from all stores")
            qs = qs.all()
            print(f"[STEP 2a] Total products in system: {qs.count()}")
        
        # Apply search and category filters (same as before)
        if query:
            print(f"\n[STEP 3] Applying search filter for: '{query}'")
            old_count = qs.count()
            qs = qs.filter(
                Q(name__icontains=query) | 
                Q(sku__icontains=query) | 
                Q(barcode__icontains=query)
            )
            print(f"[STEP 3a] Products found: {qs.count()} (was {old_count})")
        else:
            print(f"\n[STEP 3] No search query provided")
        
        if category:
            print(f"\n[STEP 4] Applying category filter for ID: '{category}'")
            old_count = qs.count()
            qs = qs.filter(category_id=category)
            print(f"[STEP 4a] Products in category: {qs.count()} (was {old_count})")
        else:
            print(f"\n[STEP 4] No category filter provided")
        
        # Final results
        print(f"\n[STEP 5] FINAL RESULTS")
        print(f"[STEP 5a] Total products after all filters: {qs.count()}")
        
        final_qs = qs[:50]
        print(f"[STEP 5b] Returning {len(final_qs)} products")
        
        # Print detailed results
        print(f"\n[FINAL PRODUCTS LIST]:")
        current_store = None
        if force_store_filter:
            try:
                current_store = get_current_store(request)
            except:
                pass
        
        total_stock_sum = 0        
        for idx, product in enumerate(final_qs, 1):
            if current_store:
                stock_item = StoreStock.objects.filter(store=current_store, product=product).first()
                stock_qty = stock_item.quantity if stock_item else 0
                total_stock_sum += stock_qty
                print(f"  {idx}. ID:{product.id} | {product.name} | Stock in current store: {stock_qty}")
            else:
                total_stock = StoreStock.objects.filter(product=product).aggregate(total=models.Sum('quantity'))['total'] or 0
                print(f"  {idx}. ID:{product.id} | {product.name} | Total stock across all stores: {total_stock}")
        
        print(f"\n[SUMMARY] Total unique products: {len(final_qs)} | Total items in current store: {total_stock_sum}")
        
        serializer = self.get_serializer(final_qs, many=True)
        
        response_data = {
            "context": "pos" if is_pos_context else context,
            "store_filtered": force_store_filter,
            "show_only_in_stock": show_only_in_stock,
            "count": len(final_qs),
            "total_items_in_store": total_stock_sum,  # ← ADD THIS: Shows sum of all stock
            "results": serializer.data
        }
        
        if current_store:
            response_data["store"] = {
                "id": current_store.id,
                "name": current_store.name
            }
        
        print("="*50 + "\n")
        
        return Response(response_data)


    @action(detail=True, methods=["get"], url_path="check_stock")
    def check_stock(self, request, pk=None):
        """Check if product has sufficient stock in current store for checkout"""
        product = self.get_object()
        
        try:
            store = get_current_store(request)
            if not store:
                return Response(
                    {"error": "No store selected. Please select a store before checkout."}, 
                    status=400
                )
                
            quantity_needed = int(request.query_params.get('quantity', 0))
            
            store_stock = StoreStock.objects.filter(
                product=product, 
                store=store
            ).first()
            
            current_stock = store_stock.quantity if store_stock else 0
            has_sufficient = current_stock >= quantity_needed
            
            return Response({
                "success": True,
                "product_id": product.id,
                "product_name": product.name,
                "product_sku": product.sku,
                "store_id": store.id,
                "store_name": store.name,
                "current_stock": current_stock,
                "quantity_needed": quantity_needed,
                "has_sufficient_stock": has_sufficient,
                "missing_quantity": max(0, quantity_needed - current_stock),
                "can_checkout": has_sufficient
            })
            
        except PermissionDenied as e:
            return Response({"error": str(e)}, status=403)
        except ValueError:
            return Response({"error": "Invalid quantity parameter"}, status=400)

    # @action(detail=True, methods=["post"], url_path="update_store_stock")
    # def update_store_stock(self, request, pk=None):
    #     product = self.get_object()
        
    #     # Get store from request or session
    #     store_id = request.data.get("store_id")
        
    #     if not store_id:
    #         # Try to get from session if not provided
    #         try:
    #             store = get_current_store(request)
    #             if store:
    #                 store_id = store.id
    #             else:
    #                 return Response({"error": "store_id is required or select a store"}, status=400)
    #         except PermissionDenied as e:
    #             return Response({"error": str(e)}, status=403)

    #     action_type = request.data.get("action", "set")

    #     try:
    #         quantity = int(request.data.get("quantity"))
    #     except (TypeError, ValueError):
    #         return Response({"error": "Invalid quantity"}, status=400)

    #     # get or create stock row
    #     store_stock, created = StoreStock.objects.get_or_create(
    #         product=product,
    #         store_id=store_id,
    #         defaults={"quantity": 0}
    #     )

    #     if action_type == "set":
    #         store_stock.quantity = quantity
    #         store_stock.save()

    #     elif action_type == "add":
    #         StoreStock.objects.filter(id=store_stock.id).update(
    #             quantity=F("quantity") + quantity
    #         )

    #     elif action_type == "remove":
    #         with transaction.atomic():
    #             stock = StoreStock.objects.select_for_update().get(id=store_stock.id)

    #             if stock.quantity < quantity:
    #                 return Response({"error": "Insufficient stock"}, status=400)

    #             stock.quantity -= quantity
    #             stock.save()

    #     else:
    #         return Response({"error": "Invalid action. Use 'set', 'add', or 'remove'"}, status=400)

    #     store_stock.refresh_from_db()

    #     return Response({
    #         "message": "Stock updated successfully",
    #         "store_id": store_id,
    #         "product_id": product.id,
    #         "quantity": store_stock.quantity
    #     })



    @action(detail=True, methods=["post"], url_path="update_store_stock")
    def update_store_stock(self, request, pk=None):
        product = self.get_object()

        store_id = request.data.get("store_id")

        if not store_id:
            try:
                store = get_current_store(request)

                if store:
                    store_id = store.id
                else:
                    return Response(
                        {"error": "store_id is required or select a store"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            except PermissionDenied as e:
                return Response(
                    {"error": str(e)},
                    status=status.HTTP_403_FORBIDDEN
                )

        try:
            quantity = int(request.data.get("quantity"))
        except (TypeError, ValueError):
            return Response(
                {"error": "Invalid quantity"},
                status=status.HTTP_400_BAD_REQUEST
            )

        action_type = request.data.get("action", "set")

        try:
            store = Store.objects.get(id=store_id)

            # if action_type == "set":
            #     StoreStockService.set(
            #         product=product,
            #         store=store,
            #         quantity=quantity
            #     )

            # elif action_type == "add":
            #     StoreStockService.add(
            #         product=product,
            #         store=store,
            #         quantity=quantity,
            #         user=request.user,
            #         remarks="Manual stock addition"
            #     )

            # elif action_type == "subtract":
            #     StoreStockService.subtract(
            #         product=product,
            #         store=store,
            #         quantity=quantity,
            #         user=request.user,
            #         remarks="Manual stock removal"
            #     )
            if action_type == "set":
                StoreStockService.adjust_stock(
                    product=product,
                    store=store,
                    action="set",
                    quantity=quantity,
                    user=request.user,
                    remarks="Manual stock set"
                )

            elif action_type == "add":
                StoreStockService.adjust_stock(
                    product=product,
                    store=store,
                    action="add",
                    quantity=quantity,
                    user=request.user,
                    remarks="Manual stock addition"
                )

            elif action_type == "subtract":
                StoreStockService.adjust_stock(
                    product=product,
                    store=store,
                    action="subtract",
                    quantity=quantity,
                    user=request.user,
                    remarks="Manual stock removal"
                )
            else:
                return Response(
                    {"error": "Invalid action. Use 'set', 'add', or 'subtract'"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            current_quantity = StoreStockService.get_store_stock(product, store)

            return Response({
                "message": "Stock updated successfully",
                "store_id": store.id,
                "product_id": product.id,
                "quantity": current_quantity
            })

        except Store.DoesNotExist:
            return Response(
                {"error": "Store not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        except ValidationError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=["post"], url_path="toggle_active")
    def toggle_active(self, request, pk=None):
        print("\n[PRODUCT API] toggle_active called")
        product = self.get_object()
        old_status = product.is_active
        product.is_active = not old_status
        product.save()
        print("[PRODUCT API] is_active:", old_status, "→", product.is_active)
        return Response({"is_active": product.is_active})


    @action(detail=False, methods=["get"], url_path="store_products_summary")
    def store_products_summary(self, request):
        """Get a summary of products across all stores (admin only)"""
        user = request.user
        
        # Check if user is admin
        if not (user.is_superuser or user.role in ["admin", "manager"]):
            return Response({"error": "Permission denied"}, status=403)
        
        from django.db.models import Sum, Count, Avg
        
        # Get all stores
        stores = Store.objects.all()
        
        summary = []
        for store in stores:
            store_summary = {
                "store_id": store.id,
                "store_name": store.name,
                "total_products": StoreStock.objects.filter(store=store).count(),
                "total_stock_value": StoreStock.objects.filter(store=store).aggregate(
                    total=models.Sum(models.F('quantity') * models.F('product__selling_price'))
                )['total'] or 0,
                "low_stock_count": StoreStock.objects.filter(
                    store=store,
                    quantity__lte=models.F('product__reorder_level')
                ).count(),
                "out_of_stock_count": StoreStock.objects.filter(store=store, quantity=0).count(),
                "in_stock_count": StoreStock.objects.filter(store=store, quantity__gt=0).count()
            }
            summary.append(store_summary)
        
        return Response({
            "total_stores": stores.count(),
            "stores": summary
        })



@require_http_methods(["GET"])
def get_product_by_barcode(request, barcode):
    """
    API endpoint to fetch product by barcode
    """
    try:
        # Search by barcode field
        product = get_object_or_404(Product, barcode=barcode, is_active=True)
        
        # Get stock information
        store = None
        if request.user.is_authenticated and hasattr(request.user, 'current_store'):
            store = request.user.current_store
        
        response_data = {
            'id': product.id,
            'sku': product.sku,
            'barcode': product.barcode,
            'name': product.name,
            'category_id': product.category_id,
            'category_name': product.category.name if product.category else None,
            'description': product.description,
            'selling_price': float(product.selling_price),
            'cost_price': float(product.cost_price),
            'available_stock': product.available_stock,
            'warehouse_stock': product.warehouse_stock,
            'reorder_level': product.reorder_level,
            'is_taxable': product.is_taxable,
            'tax_rate': float(product.tax_rate),
            'fixed_price': product.fixed_price,
        }
        
        # Add store-specific stock if available
        if store:
            response_data['store_stock'] = product.get_store_stock(store)
        
        return JsonResponse(response_data)
        
    except Product.DoesNotExist:
        return JsonResponse({
            'error': f'Product with barcode {barcode} not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

@require_http_methods(["POST"])
def bulk_barcode_scan(request):
    """
    Handle multiple barcode scans at once
    """
    try:
        data = json.loads(request.body)
        barcodes = data.get('barcodes', [])
        
        products = []
        not_found = []
        
        for barcode in barcodes:
            try:
                product = Product.objects.get(barcode=barcode, is_active=True)
                products.append({
                    'id': product.id,
                    'name': product.name,
                    'barcode': product.barcode,
                    'selling_price': float(product.selling_price),
                    'available_stock': product.available_stock,
                })
            except Product.DoesNotExist:
                not_found.append(barcode)
        
        return JsonResponse({
            'success': True,
            'products': products,
            'not_found': not_found,
            'count': len(products)
        })
        
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

def print_product_barcode(request, product_id):
    """Print single product barcode label"""

    product = get_object_or_404(Product, id=product_id, is_active=True)

    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="barcode_{product.sku}.pdf"'

    # Label size: 60mm x 30mm
    p = canvas.Canvas(response, pagesize=(60 * mm, 30 * mm))

    # Generate barcode (EAN13 requires 12/13-digit numeric string)
    EAN13 = barcode.get_barcode_class('ean13')

    ean = EAN13(str(product.barcode), writer=ImageWriter())

    buffer = io.BytesIO()
    ean.write(buffer)
    buffer.seek(0)

    # Convert buffer into something ReportLab can use
    image = ImageReader(buffer)

    # Draw product name
    p.setFont("Helvetica", 8)
    p.drawString(5 * mm, 25 * mm, product.name[:25])

    # Draw barcode image
    p.drawImage(image, 5 * mm, 10 * mm, width=50 * mm, height=12 * mm, preserveAspectRatio=True, mask='auto')

    # Draw price
    p.setFont("Helvetica", 10)
    p.drawString(5 * mm, 5 * mm, f"${product.selling_price}")

    p.showPage()
    p.save()

    return response



 
# def print_multiple_barcodes(request):
#     """Print barcodes for multiple products"""

#     product_ids = request.GET.get('ids', '')
#     product_ids = [pid for pid in product_ids.split(',') if pid]

#     products = Product.objects.filter(id__in=product_ids, is_active=True)

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'inline; filename="barcodes.pdf"'

#     p = canvas.Canvas(response, pagesize=(210 * mm, 297 * mm))  # A4

#     x_pos = 10 * mm
#     y_pos = 280 * mm
#     labels_per_row = 3

#     for i, product in enumerate(products):

#         # Generate barcode
#         EAN13 = barcode.get_barcode_class('ean13')
#         ean = EAN13(str(product.barcode), writer=ImageWriter())

#         buffer = io.BytesIO()
#         ean.write(buffer)
#         buffer.seek(0)

#         # FIX: convert to ImageReader (important)
#         image = ImageReader(buffer)

#         # Draw product name
#         p.setFont("Helvetica", 8)
#         p.drawString(x_pos, y_pos + 20 * mm, product.name[:20])

#         # Draw barcode image
#         p.drawImage(
#             image,
#             x_pos,
#             y_pos + 5 * mm,
#             width=50 * mm,
#             height=15 * mm,
#             preserveAspectRatio=True,
#             mask='auto'
#         )

#         # Draw price
#         p.setFont("Helvetica", 10)
#         p.drawString(x_pos, y_pos, f"${product.selling_price}")

#         # Move position
#         x_pos += 65 * mm

#         if (i + 1) % labels_per_row == 0:
#             x_pos = 10 * mm
#             y_pos -= 40 * mm

#         if y_pos < 20 * mm:
#             p.showPage()
#             y_pos = 280 * mm
#             x_pos = 10 * mm

#     p.save()
#     return response

# import io

# from django.http import HttpResponse
# from reportlab.pdfgen import canvas
# from reportlab.lib.units import mm
# from reportlab.lib.utils import ImageReader

# import barcode
# from barcode.writer import ImageWriter
# from PIL import Image

# from .models import Product


def print_multiple_barcodes(request):
    """Print barcodes for multiple products"""

    product_ids = request.GET.get('ids', '')
    product_ids = [pid for pid in product_ids.split(',') if pid]

    products = Product.objects.filter(
        id__in=product_ids,
        is_active=True
    )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="barcodes.pdf"'

    # A4 page size
    p = canvas.Canvas(response, pagesize=(210 * mm, 297 * mm))

    x_pos = 10 * mm
    y_pos = 280 * mm

    labels_per_row = 3

    for i, product in enumerate(products):

        try:
            # Ensure barcode is 12 digits for EAN13
            barcode_number = str(product.barcode).zfill(12)[:12]

            # Generate barcode
            EAN13 = barcode.get_barcode_class('ean13')
            ean = EAN13(barcode_number, writer=ImageWriter())

            # Create image buffer
            buffer = io.BytesIO()

            # Save as PNG
            ean.write(buffer, options={"format": "PNG"})
            buffer.seek(0)

            # Open image with PIL
            pil_image = Image.open(buffer)

            # Convert to ReportLab image
            image = ImageReader(pil_image)

            # Draw product name
            p.setFont("Helvetica", 8)
            p.drawString(
                x_pos,
                y_pos + 20 * mm,
                product.name[:20]
            )

            # Draw barcode image
            p.drawImage(
                image,
                x_pos,
                y_pos + 5 * mm,
                width=50 * mm,
                height=15 * mm,
                preserveAspectRatio=True,
                mask='auto'
            )

            # Draw price
            p.setFont("Helvetica", 10)
            p.drawString(
                x_pos,
                y_pos,
                f"${product.selling_price}"
            )

            # Move to next label position
            x_pos += 65 * mm

            # New row
            if (i + 1) % labels_per_row == 0:
                x_pos = 10 * mm
                y_pos -= 40 * mm

            # New page
            if y_pos < 20 * mm:
                p.showPage()
                y_pos = 280 * mm
                x_pos = 10 * mm

        except Exception as e:
            print(f"Barcode generation error for product {product.id}: {e}")

    p.save()

    return response
    

def barcode_print_page(request):
    """HTML page for selecting products to print"""
    products = Product.objects.filter(is_active=True)
    return render(request, 'inventory/barcode_print_page.html', {'products': products})

def print_barcode_from_barcode(request, barcode):
    """Print barcode label directly from barcode number"""
    product = get_object_or_404(Product, barcode=barcode, is_active=True)
    return print_product_barcode(request, product.id)

    



# stote 
# @method_decorator([login_required], name='dispatch')
class StoreViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Store.objects.all()
    serializer_class = StoreSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'location']
    ordering_fields = ['name', 'store_type']
    ordering = ['name']

    @action(detail=False, methods=['get'])
    def retail_stores(self, request):
        """Get all retail stores"""
        retail_stores = self.get_queryset().filter(store_type=Store.RETAIL)
        serializer = self.get_serializer(retail_stores, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def warehouses(self, request):
        """Get all warehouses"""
        warehouses = self.get_queryset().filter(store_type=Store.WAREHOUSE)
        serializer = self.get_serializer(warehouses, many=True)
        return Response(serializer.data)



class StoreStockViewSet(viewsets.ModelViewSet):
    """
    Store stock management API (service-layer driven)
    """
    permission_classes = [IsAuthenticated]

    queryset = StoreStock.objects.all().select_related('store', 'product')
    serializer_class = StoreStockSerializer

    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['store', 'product']
    search_fields = ['product__name', 'product__sku', 'store__name']

    def partial_update(self, request, *args, **kwargs):
        """
        Handles stock adjustment via service layer
        JS calls: PATCH /store-stocks/{id}/
        """
        try:
            stock = self.get_object()
            
            # Debug logging
            print(f"Received PATCH request for stock {stock.id}")
            print(f"Request data: {request.data}")
            
            action = request.data.get("action")
            quantity = request.data.get("quantity")
            notes = request.data.get("notes", "")
            
            print(f"Action: {action}, Quantity: {quantity}, Notes: {notes}")
            
            # Validate required fields
            if not action:
                return Response(
                    {"error": "action is required (set, add, or subtract)"},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            if quantity is None:
                return Response(
                    {"error": "quantity is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Convert quantity to int
            try:
                quantity_int = int(quantity)
            except (ValueError, TypeError):
                return Response(
                    {"error": "quantity must be a valid integer"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate action
            if action not in ['set', 'add', 'subtract']:
                return Response(
                    {"error": f"Invalid action: {action}. Must be 'set', 'add', or 'subtract'"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Call service layer
            # updated_stock = StoreStockService.adjust_stock(
            #     stock=stock,
            #     action=action,
            #     quantity=quantity_int,
            #     user=request.user,
            #     notes=notes,
            # )
            StoreStockService.adjust_stock(
                product=stock.product,
                store=stock.store,
                action=action,
                quantity=quantity_int,
                user=request.user,
                remarks=notes,
            )

            stock.refresh_from_db()

            serializer = self.get_serializer(stock)
            return Response(serializer.data)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            print(f"Unexpected error: {str(e)}")
            return Response(
                {"error": f"An unexpected error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    # =========================================================
    # STORE INVENTORY
    # =========================================================
    @action(detail=False, methods=['get'])
    def store_inventory(self, request):
        """
        Get complete inventory for a specific store
        """
        store_id = request.query_params.get('store_id')

        if not store_id:
            return Response(
                {'error': 'store_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        store_stocks = self.get_queryset().filter(store_id=store_id)
        serializer = self.get_serializer(store_stocks, many=True)
        return Response(serializer.data)

    # =========================================================
    # LOW STOCK ALERTS
    # =========================================================
    @action(detail=False, methods=['get'])
    def low_stock_alerts(self, request):
        """
        Get low stock alerts across all stores
        """
        low_stock = self.get_queryset().filter(
            quantity__lte=F('product__reorder_level')
        ).select_related('product', 'store')

        serializer = self.get_serializer(low_stock, many=True)
        return Response(serializer.data)

    # =========================================================
    # RETAIL STOCK (FOR POS SYSTEM)
    # =========================================================
    @action(detail=False, methods=['get'])
    def retail_stock(self, request):
        """
        Get stock only from retail stores (for POS)
        """
        retail_stocks = self.get_queryset().filter(
            store__store_type=Store.RETAIL
        )

        serializer = self.get_serializer(retail_stocks, many=True)
        return Response(serializer.data)


@method_decorator(csrf_exempt, name='dispatch')
class StockTransactionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = StockTransactionSerializer
    queryset = StockTransaction.objects.all().select_related(
        'product', 'store', 'transfer_to_store', 'performed_by'
    )

    # Filters for list endpoint
    def get_queryset(self):
        queryset = super().get_queryset()

        transaction_type = self.request.query_params.get('transaction_type')
        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)

        product_id = self.request.query_params.get('product')
        if product_id:
            queryset = queryset.filter(product_id=product_id)

        store_id = self.request.query_params.get('store')
        if store_id:
            queryset = queryset.filter(store_id=store_id)

        date_from = self.request.query_params.get('date_from')
        if date_from:
            queryset = queryset.filter(timestamp__gte=date_from)

        date_to = self.request.query_params.get('date_to')
        if date_to:
            queryset = queryset.filter(timestamp__lte=date_to)

        return queryset.order_by('-timestamp')

    # Use request.user directly (simpler and more reliable)
    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        with transaction.atomic():
            stock_transaction = serializer.save(performed_by=user)
            self.update_stock_levels(stock_transaction)

    # Update stock levels based on transaction type
    def update_stock_levels(self, transaction_obj):
        try:
            if transaction_obj.transaction_type == 'IN':
                store_stock, created = StoreStock.objects.get_or_create(
                    store=transaction_obj.store,
                    product=transaction_obj.product,
                    defaults={'quantity': transaction_obj.quantity}
                )
                if not created:
                    store_stock.quantity += transaction_obj.quantity
                    store_stock.save()

            elif transaction_obj.transaction_type == 'OUT':
                store_stock = StoreStock.objects.get(
                    store=transaction_obj.store,
                    product=transaction_obj.product
                )
                if store_stock.quantity < transaction_obj.quantity:
                    raise ValueError(
                        f"Insufficient stock. Available: {store_stock.quantity}, Requested: {transaction_obj.quantity}"
                    )
                store_stock.quantity -= transaction_obj.quantity
                store_stock.save()

            elif transaction_obj.transaction_type == 'TRANSFER':
                if not transaction_obj.transfer_to_store:
                    raise ValueError("Transfer destination store is required")

                # Source store
                source_stock = StoreStock.objects.get(
                    store=transaction_obj.store,
                    product=transaction_obj.product
                )
                if source_stock.quantity < transaction_obj.quantity:
                    raise ValueError(
                        f"Insufficient stock in source store. Available: {source_stock.quantity}, Requested: {transaction_obj.quantity}"
                    )
                source_stock.quantity -= transaction_obj.quantity
                source_stock.save()

                # Destination store
                dest_stock, created = StoreStock.objects.get_or_create(
                    store=transaction_obj.transfer_to_store,
                    product=transaction_obj.product,
                    defaults={'quantity': transaction_obj.quantity}
                )
                if not created:
                    dest_stock.quantity += transaction_obj.quantity
                    dest_stock.save()

        except ObjectDoesNotExist as e:
            raise ValueError(f"Stock record not found: {str(e)}")
        except Exception as e:
            raise ValueError(f"Error updating stock levels: {str(e)}")

    # Override create() to return proper error messages
    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"Unexpected error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Summary endpoint
    @action(detail=False, methods=['get'])
    def summary(self, request):
        queryset = self.get_queryset()

        summary = {
            'in': {
                'count': queryset.filter(transaction_type='IN').count(),
                'quantity': sum(t.quantity for t in queryset.filter(transaction_type='IN'))
            },
            'out': {
                'count': queryset.filter(transaction_type='OUT').count(),
                'quantity': sum(t.quantity for t in queryset.filter(transaction_type='OUT'))
            },
            'transfer': {
                'count': queryset.filter(transaction_type='TRANSFER').count(),
                'quantity': sum(t.quantity for t in queryset.filter(transaction_type='TRANSFER'))
            }
        }

        summary['total'] = {
            'count': sum(item['count'] for item in summary.values()),
            'quantity': sum(item['quantity'] for item in summary.values())
        }

        return Response(summary)


@method_decorator(csrf_exempt, name='dispatch')
class StockTransferViewSet(viewsets.ModelViewSet):

    queryset = StockTransfer.objects.all().select_related(
        'product',
        'from_store',
        'to_store',
        'performed_by'
    )

    filter_backends = [
        DjangoFilterBackend,
        filters.OrderingFilter,
        filters.SearchFilter
    ]

    filterset_fields = [
        'from_store',
        'to_store',
        'product',
        'status'
    ]

    search_fields = [
        'product__name',
        'product__sku',
        'notes',
        'transfer_code'
    ]

    ordering_fields = [
        'created_at',
        'quantity'
    ]

    ordering = ['-created_at']

    # -------------------------------------------------
    # DYNAMIC SERIALIZER
    # -------------------------------------------------
    def get_serializer_class(self):

        if self.action == 'create':
            return StockTransferCreateSerializer

        return StockTransferSerializer

    # -------------------------------------------------
    # QUERYSET FILTERING
    # -------------------------------------------------
    def get_queryset(self):

        queryset = super().get_queryset()

        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        store_type = self.request.query_params.get('store_type')

        if date_from:
            queryset = queryset.filter(
                created_at__date__gte=date_from
            )

        if date_to:
            queryset = queryset.filter(
                created_at__date__lte=date_to
            )

        if store_type:
            queryset = queryset.filter(
                Q(from_store__store_type=store_type) |
                Q(to_store__store_type=store_type)
            )

        return queryset

    # -------------------------------------------------
    # CREATE TRANSFER
    # -------------------------------------------------
    

    # def create(self, request, *args, **kwargs):

    #     serializer = self.get_serializer(data=request.data)
    #     serializer.is_valid(raise_exception=True)

    #     validated = serializer.validated_data

    #     user = (
    #         request.user
    #         if request.user.is_authenticated
    #         else None
    #     )

    #     result = StoreStockService.transfer_stock(
    #         product=validated['product'],
    #         from_store=validated['from_store'],
    #         to_store=validated['to_store'],
    #         quantity=validated['quantity'],
    #         user=user,
    #         reference=f"TRANSFER-{timezone.now().strftime('%Y%m%d%H%M%S')}",
    #         remarks=validated.get('notes', '')
    #     )
    #     print("the ")
    #     response_serializer = StockTransferSerializer(
    #         result["transfer"],
    #         context={'request': request}
    #     )

    #     return Response(
    #         response_serializer.data,
    #         status=status.HTTP_201_CREATED
    #     )
    
    # -------------------------------------------------
    # CREATE TRANSFER
    # -------------------------------------------------
    def create(self, request, *args, **kwargs):

        print("\n========== STOCK TRANSFER REQUEST ==========")
        print("Incoming request data:", request.data)

        serializer = self.get_serializer(data=request.data)

        if not serializer.is_valid():

            print("\n❌ VALIDATION FAILED")
            print("Errors:", serializer.errors)

            return Response(
                serializer.errors,
                status=status.HTTP_400_BAD_REQUEST
            )

        print("\n✅ VALIDATION PASSED")

        validated = serializer.validated_data

        user = (
            request.user
            if request.user.is_authenticated
            else None
        )

        print("\n--- TRANSFER DETAILS ---")
        print("Product:", validated['product'])
        print("From Store:", validated['from_store'])
        print("To Store:", validated['to_store'])
        print("Quantity:", validated['quantity'])
        print("User:", user)
        print("Notes:", validated.get('notes', ''))

        try:

            print("\n🚀 STARTING TRANSFER SERVICE")

            result = StoreStockService.transfer_stock(
                product=validated['product'],
                from_store=validated['from_store'],
                to_store=validated['to_store'],
                quantity=validated['quantity'],
                user=user,
                reference=f"TRANSFER-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                remarks=validated.get('notes', '')
            )

            print("\n✅ TRANSFER SERVICE COMPLETED")

            print("\n--- TRANSFER RESULT ---")
            print("Transfer Code:", result["transfer"].transfer_code)
            print("Transfer Object:", result["transfer"])
            print("OUT Transaction:", result["out_transaction"])
            print("IN Transaction:", result["in_transaction"])

        except Exception as e:

            print("\n❌ TRANSFER FAILED")
            print("Error:", str(e))

            return Response(
                {
                    "detail": str(e)
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        response_serializer = StockTransferSerializer(
            result["transfer"],
            context={'request': request}
        )

        print("\n✅ RESPONSE SERIALIZED")
        print("Response Data:", response_serializer.data)

        print("\n========== TRANSFER COMPLETE ==========\n")

        return Response(
            response_serializer.data,
            status=status.HTTP_201_CREATED
        )
    # -------------------------------------------------
    # SUMMARY
    # -------------------------------------------------
    @action(detail=False, methods=['get'])
    def summary(self, request):

        queryset = self.filter_queryset(
            self.get_queryset()
        )

        summary = {
            'total_transfers': queryset.count(),

            'total_quantity': queryset.aggregate(
                total=Sum('quantity')
            )['total'] or 0,

            'today': queryset.filter(
                created_at__date=timezone.now().date()
            ).count(),

            'by_store': {}
        }

        stores = set()

        for transfer in queryset:
            stores.add(transfer.from_store)
            stores.add(transfer.to_store)

        for store in stores:

            outgoing = queryset.filter(
                from_store=store
            ).aggregate(
                total=Sum('quantity')
            )['total'] or 0

            incoming = queryset.filter(
                to_store=store
            ).aggregate(
                total=Sum('quantity')
            )['total'] or 0

            summary['by_store'][store.name] = {
                'outgoing': outgoing,
                'incoming': incoming,
                'net': incoming - outgoing
            }

        return Response(summary)

    # -------------------------------------------------
    # RECENT TRANSFERS
    # -------------------------------------------------
    @action(detail=False, methods=['get'])
    def recent(self, request):

        recent_date = timezone.now() - timezone.timedelta(days=7)

        recent_transfers = self.get_queryset().filter(
            created_at__gte=recent_date
        )[:10]

        serializer = self.get_serializer(
            recent_transfers,
            many=True
        )

        return Response(serializer.data)




# @login_required
def inventory_summary(request):
    total_products = Product.objects.count()
    total_categories = Category.objects.count()
    total_stores = Store.objects.count()
    
    # Low stock in retail stores
    retail_stores = Store.objects.filter(store_type=Store.RETAIL)
    low_stock_count = StoreStock.objects.filter(
        store__in=retail_stores,
        quantity__lte=F('product__reorder_level')
    ).count()
    
    # Out of stock in retail stores
    out_of_stock_count = StoreStock.objects.filter(
        store__in=retail_stores,
        quantity=0
    ).count()
    
    context = {
        'total_products': total_products,
        'total_categories': total_categories,
        'total_stores': total_stores,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
    }
    
    return render(request, 'inventory/inventory_summary.html', context)


@method_decorator(csrf_exempt, name='dispatch')
class BulkRestockXlsViewSet(viewsets.ModelViewSet):
    serializer_class = BulkRestockSerializer
    queryset = BulkRestock.objects.all().select_related('store', 'category')

    # ---------------------------------------------------
    # 1. HTML PAGE ACTION (Loads bulk_restock.html)
    # ---------------------------------------------------
    @action(detail=False, methods=['get'])
    def page(self, request):
        context = {
            "stores": Store.objects.all(),
            "categories": Category.objects.all(),
        }
        # return render(request, 'inventory/bulk_restock.html', context)

    # ---------------------------------------------------
    # 2. GENERATE RESTOCK EXCEL TEMPLATE
    # ---------------------------------------------------
    @action(detail=False, methods=['post'])
    def generate_template(self, request):
        store_id = request.data.get('store_id')
        category_id = request.data.get('category_id')
        include_all = request.data.get('include_all', False)
        low_stock_only = request.data.get('low_stock_only', False)

        if not store_id:
            return Response({'error': 'store_id is required'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            store = Store.objects.get(id=store_id)
        except Store.DoesNotExist:
            return Response({'error': 'Store not found'},
                            status=status.HTTP_404_NOT_FOUND)

        # Create restock record - ONLY with fields that exist in the model
        restock = BulkRestock.objects.create(
            store=store,
            category_id=category_id if category_id else None,
            include_all=include_all,
            notes=request.data.get('notes', '') or f"Generated via Excel template"
        )
        
        # Store low_stock_only flag in notes since model doesn't have this field
        if low_stock_only:
            if restock.notes:
                restock.notes += f" | Filter: Low Stock Only"
            else:
                restock.notes = "Filter: Low Stock Only"
            restock.save()

        # Get products based on filters
        products = Product.objects.filter(is_active=True)
        if category_id:
            products = products.filter(category_id=category_id)

        products_data = []
        items_created = 0

        for product in products:
            # Get current stock for this store
            try:
                stock = StoreStock.objects.get(store=store, product=product)
                current_qty = stock.quantity
            except StoreStock.DoesNotExist:
                # If no stock record exists, assume 0
                current_qty = 0
                # Create a stock record with 0 quantity
                StoreStock.objects.create(store=store, product=product, quantity=0)

            # Filter low stock if requested
            if low_stock_only:
                reorder_level = product.reorder_level
                if current_qty > reorder_level:  # Changed to > instead of >=
                    continue  # Skip products that are NOT low stock

            # Create restock item
            BulkRestockItem.objects.create(
                restock=restock,
                product=product,
                current_quantity=current_qty,
                new_quantity=current_qty,  # Default to current quantity
                current_price=product.selling_price,
                new_price=product.selling_price  # Default to current price
            )
            items_created += 1

            # products_data.append({
            #     'ID': product.id,
            #     'Product': product.name,
            #     'SKU': product.sku,
            #     'Category': product.category.name if product.category else '',
            #     'Current Stock': current_qty,
            #     'Reorder Level': product.reorder_level,
            #     'New Stock': current_qty,  # User will update this
            #     'Current Price': float(product.selling_price),
            #     'New Price': float(product.selling_price),  # User can update this
            #     'Restock ID': restock.id  # For reference
            # })

            products_data.append({
                'Product': product.name,
                'Category': product.category.name if product.category else '',
                'Current Stock': current_qty,
                'Current Price': float(product.selling_price),
                'New Stock': '',
                'New Price': ''
            })

        if not products_data:
            restock.delete()
            return Response({'error': 'No products match your criteria'},
                            status=status.HTTP_400_BAD_REQUEST)

        # Create Excel file
        df = pd.DataFrame(products_data)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main template sheet
            df.to_excel(writer, sheet_name='Restock Template', index=False)
            
            # Auto-adjust column widths
            ws = writer.sheets['Restock Template']
            for col in df.columns:
                col_idx = df.columns.get_loc(col)
                col_letter = chr(65 + col_idx)
                max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
                ws.column_dimensions[col_letter].width = min(max_length, 50)
            
            # Add formatting for user-editable columns
            for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                # Make New Stock (G) and New Price (H) columns light green
                ws[f'G{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                ws[f'H{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
          
            # Add instructions sheet
            instruction_data = {
                'Column': [
                    'ID',
                    'Product', 
                    'SKU',
                    'Category',
                    'Current Stock',
                    'Reorder Level',
                    'New Stock',
                    'Current Price',
                    'New Price'
                ],
                'Description': [
                    'DO NOT MODIFY - Unique product identifier',
                    'DO NOT MODIFY - Product name',
                    'DO NOT MODIFY - Stock Keeping Unit',
                    'DO NOT MODIFY - Product category',
                    'DO NOT MODIFY - Current stock quantity in store',
                    'DO NOT MODIFY - Reorder level for this product',
                    'UPDATE THIS - Enter new desired stock quantity',
                    'DO NOT MODIFY - Current selling price',
                    'OPTIONAL - Enter new price if changing'
                ],
                'Format': [
                    'Number (read-only)',
                    'Text (read-only)',
                    'Text (read-only)',
                    'Text (read-only)',
                    'Number (read-only)',
                    'Number (read-only)',
                    'Whole number (≥ 0)',
                    'Decimal (read-only)',
                    'Decimal (≥ 0)'
                ]
            }
            instruction_df = pd.DataFrame(instruction_data)
            instruction_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Add summary sheet
            summary_data = {
                'Field': [
                    'Store',
                    'Category Filter',
                    'Stock Filter',
                    'Template Generated',
                    'Total Products',
                    'Restock ID',
                    'Instructions'
                ],
                'Value': [
                    store.name,
                    Category.objects.get(id=category_id).name if category_id else 'All Categories',
                    'Low Stock Only' if low_stock_only else 'All Products',
                    timezone.now().strftime('%Y-%m-%d %H:%M'),
                    len(products_data),
                    restock.id,
                    'Fill in "New Stock" column and upload back to complete restock'
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

        output.seek(0)

        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="restock_{store.name}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        )
        # Add restock ID to header for client-side storage
        response['X-Restock-ID'] = str(restock.id)
        
        return response

    # ---------------------------------------------------
    # 3. UPLOAD COMPLETED TEMPLATE
    # ---------------------------------------------------
    # @action(detail=True, methods=['post'])
    # def upload_completed(self, request, pk=None):
    #     print("the upload is ", request)
    #     restock = self.get_object()

    #     if restock.completed:
    #         return Response({'error': 'Restock already completed'},
    #                         status=status.HTTP_400_BAD_REQUEST)

    #     file = request.FILES.get("file")
    #     if not file:
    #         return Response({'error': 'Excel file is required'},
    #                         status=status.HTTP_400_BAD_REQUEST)

    #     try:
    #         # Read Excel file
    #         df = pd.read_excel(file, sheet_name='Restock Template')
            
    #         # Validate required columns
    #         required_columns = ['ID', 'New Stock']
    #         missing_columns = [col for col in required_columns if col not in df.columns]
    #         if missing_columns:
    #             return Response({'error': f'Missing required columns: {", ".join(missing_columns)}'},
    #                             status=status.HTTP_400_BAD_REQUEST)

    #         updated_count = 0
    #         errors = []
    #         error_rows = []

    #         with transaction.atomic():
    #             for index, row in df.iterrows():
    #                 row_num = index + 2  # Excel rows start at 1, plus header row
                    
    #                 # Skip empty rows
    #                 if pd.isna(row.get('ID')):
    #                     continue
                    
    #                 try:
    #                     product_id = int(row['ID'])
    #                 except (ValueError, TypeError):
    #                     error_msg = f"Invalid Product ID: {row.get('ID')}"
    #                     errors.append(f"Row {row_num}: {error_msg}")
    #                     error_rows.append({
    #                         'Row': row_num,
    #                         'ID': row.get('ID'),
    #                         'Product': row.get('Product', ''),
    #                         'Error': error_msg
    #                     })
    #                     continue

    #                 try:
    #                     item = BulkRestockItem.objects.get(
    #                         restock=restock,
    #                         product_id=product_id
    #                     )
    #                 except BulkRestockItem.DoesNotExist:
    #                     error_msg = f"Product ID {product_id} not found in this restock session"
    #                     errors.append(f"Row {row_num}: {error_msg}")
    #                     error_rows.append({
    #                         'Row': row_num,
    #                         'ID': product_id,
    #                         'Product': row.get('Product', ''),
    #                         'Error': error_msg
    #                     })
    #                     continue

    #                 # Validate and update stock quantity
    #                 try:
    #                     new_qty = float(row['New Stock'])
    #                     if new_qty < 0:
    #                         error_msg = "New Stock cannot be negative"
    #                         errors.append(f"Row {row_num}: {error_msg}")
    #                         error_rows.append({
    #                             'Row': row_num,
    #                             'ID': product_id,
    #                             'Product': row.get('Product', ''),
    #                             'Error': error_msg
    #                         })
    #                         continue
    #                     # Convert to integer if it's a whole number
    #                     if new_qty.is_integer():
    #                         new_qty = int(new_qty)
    #                 except (ValueError, TypeError):
    #                     error_msg = "Invalid New Stock value (must be a number)"
    #                     errors.append(f"Row {row_num}: {error_msg}")
    #                     error_rows.append({
    #                         'Row': row_num,
    #                         'ID': product_id,
    #                         'Product': row.get('Product', ''),
    #                         'Error': error_msg
    #                     })
    #                     continue

    #                 # Update price if provided and valid
    #                 new_price = None
    #                 if 'New Price' in df.columns and not pd.isna(row['New Price']):
    #                     try:
    #                         new_price = float(row['New Price'])
    #                         if new_price < 0:
    #                             error_msg = "New Price cannot be negative"
    #                             errors.append(f"Row {row_num}: {error_msg}")
    #                             error_rows.append({
    #                                 'Row': row_num,
    #                                 'ID': product_id,
    #                                 'Product': row.get('Product', ''),
    #                                 'Error': error_msg
    #                             })
    #                             continue
    #                     except (ValueError, TypeError):
    #                         error_msg = "Invalid New Price value (must be a number)"
    #                         errors.append(f"Row {row_num}: {error_msg}")
    #                         error_rows.append({
    #                             'Row': row_num,
    #                             'ID': product_id,
    #                             'Product': row.get('Product', ''),
    #                             'Error': error_msg
    #                         })
    #                         continue

    #                 # Update the item
    #                 item.new_quantity = new_qty
    #                 if new_price is not None:
    #                     item.new_price = new_price
    #                 item.save()
    #                 updated_count += 1

    #             # If there are errors, create error report
    #             if errors:
    #                 error_file_base64 = self._create_error_report(error_rows, errors)
                    
    #                 return Response({
    #                     'error': f'Found {len(errors)} errors. Please fix and re-upload.',
    #                     'errors': errors[:10],  # Return first 10 errors
    #                     'error_count': len(errors),
    #                     'successful_updates': updated_count,
    #                     'error_file': error_file_base64
    #                 }, status=status.HTTP_400_BAD_REQUEST)

    #             # Process the restock if no errors
    #             restock.process_restock()
    #             restock.notes = request.data.get('notes', restock.notes)
    #             restock.completed_by = request.user if request.user.is_authenticated else None
    #             restock.save()

    #         return Response({
    #             "message": f"✅ Restock completed successfully! {updated_count} items updated.",
    #             "restock_id": restock.id,
    #             "completed": True,
    #             "items_updated": updated_count,
    #             "store": restock.store.name,
    #             "timestamp": timezone.now().isoformat()
    #         })

    #     except Exception as e:
    #         return Response({'error': f'Processing error: {str(e)}'},
    #                         status=status.HTTP_400_BAD_REQUEST)
    @action(detail=True, methods=['post'])
    def upload_completed(self, request, pk=None):

        print("\n========== UPLOAD COMPLETED START ==========")
        print("REQUEST:", request)
        print("METHOD:", request.method)
        print("PATH:", request.path)
        print("CONTENT TYPE:", request.content_type)
        print("FILES:", request.FILES)
        print("DATA:", request.data)

        restock = self.get_object()

        print("RESTOCK ID:", restock.id)

        if restock.completed:
            print("ERROR: Restock already completed")
            return Response(
                {'error': 'Restock already completed'},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES.get("file")

        print("UPLOADED FILE:", file)

        if not file:
            print("ERROR: No file uploaded")
            return Response(
                {'error': 'Excel file is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            print("Reading Excel file...")

            df = pd.read_excel(file, sheet_name='Restock Template')

            print("Excel loaded successfully")
            print("COLUMNS:", list(df.columns))
            print("TOTAL ROWS:", len(df))

            # Validate required columns
            required_columns = ['ID', 'New Stock']

            missing_columns = [
                col for col in required_columns if col not in df.columns
            ]

            if missing_columns:
                print("ERROR: Missing columns:", missing_columns)

                return Response(
                    {
                        'error': f'Missing required columns: {", ".join(missing_columns)}'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            updated_count = 0
            errors = []
            error_rows = []

            # REMOVED transaction.atomic()

            for index, row in df.iterrows():

                row_num = index + 2

                print(f"\n----- PROCESSING ROW {row_num} -----")
                print("RAW ROW:", row.to_dict())

                # Skip empty rows
                if pd.isna(row.get('ID')):
                    print("Skipping empty row")
                    continue

                try:
                    product_id = int(row['ID'])
                    print("PRODUCT ID:", product_id)

                except (ValueError, TypeError) as e:

                    print("INVALID PRODUCT ID:", e)

                    error_msg = f"Invalid Product ID: {row.get('ID')}"

                    errors.append(f"Row {row_num}: {error_msg}")

                    error_rows.append({
                        'Row': row_num,
                        'ID': row.get('ID'),
                        'Product': row.get('Product', ''),
                        'Error': error_msg
                    })

                    continue

                try:
                    item = BulkRestockItem.objects.get(
                        restock=restock,
                        product_id=product_id
                    )

                    print("FOUND ITEM:", item.id)

                except BulkRestockItem.DoesNotExist:

                    print("ITEM NOT FOUND")

                    error_msg = (
                        f"Product ID {product_id} not found in this restock session"
                    )

                    errors.append(f"Row {row_num}: {error_msg}")

                    error_rows.append({
                        'Row': row_num,
                        'ID': product_id,
                        'Product': row.get('Product', ''),
                        'Error': error_msg
                    })

                    continue

                # Validate stock quantity
                try:
                    new_qty = float(row['New Stock'])

                    print("NEW QTY:", new_qty)

                    if pd.isna(new_qty):
                        raise ValueError("Quantity is NaN")

                    if new_qty < 0:

                        print("NEGATIVE QUANTITY DETECTED")

                        error_msg = "New Stock cannot be negative"

                        errors.append(f"Row {row_num}: {error_msg}")

                        error_rows.append({
                            'Row': row_num,
                            'ID': product_id,
                            'Product': row.get('Product', ''),
                            'Error': error_msg
                        })

                        continue

                    if new_qty.is_integer():
                        new_qty = int(new_qty)

                    print("FINAL NEW QTY:", new_qty)

                except (ValueError, TypeError) as e:

                    print("INVALID STOCK VALUE:", e)

                    error_msg = "Invalid New Stock value (must be a number)"

                    errors.append(f"Row {row_num}: {error_msg}")

                    error_rows.append({
                        'Row': row_num,
                        'ID': product_id,
                        'Product': row.get('Product', ''),
                        'Error': error_msg
                    })

                    continue

                # Validate price
                new_price = None

                if 'New Price' in df.columns and not pd.isna(row['New Price']):

                    try:
                        new_price = float(row['New Price'])

                        print("NEW PRICE:", new_price)

                        if new_price < 0:

                            print("NEGATIVE PRICE DETECTED")

                            error_msg = "New Price cannot be negative"

                            errors.append(f"Row {row_num}: {error_msg}")

                            error_rows.append({
                                'Row': row_num,
                                'ID': product_id,
                                'Product': row.get('Product', ''),
                                'Error': error_msg
                            })

                            continue

                    except (ValueError, TypeError) as e:

                        print("INVALID PRICE:", e)

                        error_msg = (
                            "Invalid New Price value (must be a number)"
                        )

                        errors.append(f"Row {row_num}: {error_msg}")

                        error_rows.append({
                            'Row': row_num,
                            'ID': product_id,
                            'Product': row.get('Product', ''),
                            'Error': error_msg
                        })

                        continue

                # SAVE ITEM
                try:

                    print("Saving item...")

                    item.new_quantity = new_qty

                    if new_price is not None:
                        item.new_price = new_price

                    item.save()

                    print("ITEM SAVED SUCCESSFULLY")

                    updated_count += 1

                except Exception as save_error:

                    print("ITEM SAVE FAILED")
                    print("ERROR:", str(save_error))

                    errors.append(
                        f"Row {row_num}: Save failed - {str(save_error)}"
                    )

                    error_rows.append({
                        'Row': row_num,
                        'ID': product_id,
                        'Product': row.get('Product', ''),
                        'Error': str(save_error)
                    })

            print("\n========== VALIDATION COMPLETE ==========")
            print("TOTAL ERRORS:", len(errors))
            print("UPDATED COUNT:", updated_count)

            # Generate error report
            if errors:

                print("Generating error report...")

                error_file_base64 = self._create_error_report(
                    error_rows,
                    errors
                )

                print("ERROR REPORT GENERATED")

                return Response({
                    'error': f'Found {len(errors)} errors. Please fix and re-upload.',
                    'errors': errors[:10],
                    'error_count': len(errors),
                    'successful_updates': updated_count,
                    'error_file': error_file_base64
                }, status=status.HTTP_400_BAD_REQUEST)

            # PROCESS RESTOCK
            try:

                print("\n========== PROCESSING RESTOCK ==========")

                # restock.process_restock()
                RestockDebugService(restock).process()

                print("RESTOCK PROCESSED SUCCESSFULLY")

            except Exception as process_error:

                print("RESTOCK PROCESSING FAILED")
                print("ERROR:", str(process_error))

                errors.append(str(process_error))

                error_rows.append({
                    'Row': 'SYSTEM',
                    'ID': '',
                    'Product': '',
                    'Error': str(process_error)
                })

                error_file_base64 = self._create_error_report(
                    error_rows,
                    errors
                )

                return Response({
                    'error': f'Processing failed: {str(process_error)}',
                    'errors': errors,
                    'error_count': len(errors),
                    'successful_updates': updated_count,
                    'error_file': error_file_base64
                }, status=status.HTTP_400_BAD_REQUEST)

            restock.notes = request.data.get('notes', restock.notes)

            restock.completed_by = (
                request.user if request.user.is_authenticated else None
            )

            restock.save()

            print("\n========== SUCCESS ==========")

            return Response({
                "message": f"✅ Restock completed successfully! {updated_count} items updated.",
                "restock_id": restock.id,
                "completed": True,
                "items_updated": updated_count,
                "store": restock.store.name,
                "timestamp": timezone.now().isoformat()
            })

        except Exception as e:

            import traceback

            print("\n========== FATAL ERROR ==========")

            traceback.print_exc()

            return Response(
                {'error': f'Processing error: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
   


    # ---------------------------------------------------
    # 4. QUICK RESTOCK OPTIONS
    # ---------------------------------------------------
    # @action(detail=False, methods=['get'])
    # def quick_restock_options(self, request):
    #     store_id = request.query_params.get('store_id')
    #     if not store_id:
    #         return Response({'error': 'store_id is required'},
    #                         status=status.HTTP_400_BAD_REQUEST)

    #     try:
    #         store = Store.objects.get(id=store_id)
    #     except Store.DoesNotExist:
    #         return Response({'error': 'Store not found'},
    #                         status=status.HTTP_404_NOT_FOUND)

    #     # Get categories with product counts and low stock counts
    #     categories = Category.objects.filter(is_active=True).annotate(
    #         product_count=Count('products', filter=Q(products__is_active=True)),
    #         low_stock_count=Count(
    #             'products',
    #             filter=Q(
    #                 products__is_active=True,
    #                 products__storestocks__store=store,
    #                 products__storestocks__quantity__lte=F('products__reorder_level')
    #             )
    #         )
    #     ).values('id', 'name', 'product_count', 'low_stock_count')

    #     # Get low stock products count
    #     low_stock_total = StoreStock.objects.filter(
    #         store=store,
    #         quantity__lte=F('product__reorder_level'),
    #         product__is_active=True
    #     ).count()

    #     # Get total products count for this store
    #     total_products_store = StoreStock.objects.filter(store=store, product__is_active=True).count()

    #     return Response({
    #         "store": {
    #             "id": store.id,
    #             "name": store.name
    #         },
    #         "categories": list(categories),
    #         "low_stock_products": low_stock_total,
    #         "total_products": total_products_store,
    #         "summary": {
    #             "all_products": total_products_store,
    #             "low_stock": low_stock_total,
    #             "healthy_stock": total_products_store - low_stock_total
    #         }
    #     })

    # ---------------------------------------------------
    # HELPER METHOD: Create error report
    # ---------------------------------------------------
    def _create_error_report(self, error_rows, errors):
        """Create an Excel error report and return as base64 string"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Error Report"
            
            # Write headers
            headers = ['Row', 'ID', 'Product', 'Error', 'Suggestion']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
                ws.cell(row=1, column=col_num).font = ws.cell(row=1, column=col_num).font.copy(bold=True)
            
            # Write error data
            for row_num, error_row in enumerate(error_rows, 2):
                ws.cell(row=row_num, column=1, value=error_row.get('Row', ''))
                ws.cell(row=row_num, column=2, value=error_row.get('ID', ''))
                ws.cell(row=row_num, column=3, value=error_row.get('Product', ''))
                ws.cell(row=row_num, column=4, value=error_row.get('Error', ''))
                
                # Add suggestions based on error type
                error_msg = error_row.get('Error', '')
                suggestion = ''
                if 'not found' in error_msg.lower():
                    suggestion = 'Make sure Product ID exists in the system'
                elif 'negative' in error_msg.lower():
                    suggestion = 'Enter a positive number (0 or higher)'
                elif 'invalid' in error_msg.lower():
                    if 'Stock' in error_msg:
                        suggestion = 'Enter a valid number for stock quantity'
                    elif 'Price' in error_msg:
                        suggestion = 'Enter a valid number for price (e.g., 19.99)'
                elif 'missing' in error_msg.lower():
                    suggestion = 'Fill in all required columns'
                
                ws.cell(row=row_num, column=5, value=suggestion)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet
            ws_summary = wb.create_sheet(title="Summary")
            ws_summary.append(["Error Summary", ""])
            ws_summary.append(["Total Errors", len(errors)])
            ws_summary.append(["Successful Updates", len(error_rows) - len(errors)])
            ws_summary.append(["", ""])
            ws_summary.append(["Common Issues and Solutions:", ""])
            ws_summary.append(["1. Invalid Product ID", "Check that Product IDs match the original template"])
            ws_summary.append(["2. Negative Values", "Stock and price must be 0 or positive"])
            ws_summary.append(["3. Missing Columns", "Do not delete or rename columns from the template"])
            ws_summary.append(["4. Wrong Data Type", "Stock must be whole numbers, price can be decimals"])
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Convert to base64
            return base64.b64encode(output.read()).decode('utf-8')
            
        except Exception as e:
            # Fallback: return simple error list
            import json
            return base64.b64encode(
                json.dumps({
                    "errors": errors[:20],
                    "error_count": len(errors)
                }).encode('utf-8')
            ).decode('utf-8')

    # ---------------------------------------------------
    # 5. GET RESTOCK STATUS
    # ---------------------------------------------------
    @action(detail=True, methods=['get'])
    def status(self, request, pk=None):
        """Get detailed status of a restock"""
        restock = self.get_object()
        
        items = BulkRestockItem.objects.filter(restock=restock).select_related('product')
        
        item_data = []
        total_qty_change = 0
        total_price_change = 0
        
        for item in items:
            qty_change = item.new_quantity - item.current_quantity
            price_change = (item.new_price - item.current_price) if item.new_price else 0
            
            item_data.append({
                'product_id': item.product.id,
                'product_name': item.product.name,
                'current_quantity': item.current_quantity,
                'new_quantity': item.new_quantity,
                'current_price': float(item.current_price),
                'new_price': float(item.new_price) if item.new_price else float(item.current_price),
                'quantity_change': qty_change,
                'price_change': price_change
            })
            
            total_qty_change += qty_change
            total_price_change += price_change
        
        return Response({
            'restock': {
                'id': restock.id,
                'store': restock.store.name,
                'category': restock.category.name if restock.category else 'All Categories',
                'completed': restock.completed,
                'created_at': restock.generated_at,
                'completed_at': restock.completed_at,
                'notes': restock.notes,
                'total_items': len(item_data),
                'items_changed': sum(1 for item in item_data 
                                    if item['quantity_change'] != 0 or item['price_change'] != 0),
                'total_quantity_change': total_qty_change,
                'total_price_change': total_price_change
            },
            'items': item_data
        })